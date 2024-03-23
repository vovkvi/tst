#!/usr/bin/env python3
# coding: utf-8

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.cell import MergedCell


books = []
for i in Path('./src').glob('*.xlsx'):
    if not i.is_file():
        continue
    books.append(i)

wb = load_workbook(books[0], data_only=True)

ws = wb[wb.sheetnames[0]]

res = []
for cell in ws["A"]:
    cval = ws[f'C{cell.row}'].value
    dval = ws[f'D{cell.row}'].value
    if not isinstance(cell, MergedCell):
        res.append({
            "n": str(cell.value),
            "h" : ws[f'B{cell.row}'].value,
            "o" : [f"{cval if cval is not None else ''}{dval}"]
        })
        continue
    res[-1]["o"].append(f"{cval if cval is not None else ''}{dval}")

wb.close()

with open('out.js', 'w+', encoding="utf-8", errors='ignore') as js:
    js.write(f'let data = {res};')

print('OK')
